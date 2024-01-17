import sqlite3
import csv
import os
import openpyxl
import pandas as pd
import numpy as np
import itertools
import xlwings as xw
import random


# # # Connection to the database

def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except sqlite3.Error as e:
        print(e)
    return conn




# # # Main fonction

##*## : Line to modify if you have changed the years modelled in TEMOA 

def main():
    # # ------------------------------------------------------------------------ Settings and open database file ------------------------------------------------------------------------
    Folder = os.path.realpath(__file__)[:os.path.realpath(__file__).find(os. path. basename(__file__))] # path to main directory where this script is saved
    wb_obj_in = openpyxl.load_workbook(Folder + 'run_files/' + 'Input.xlsx', data_only = True) # open the input file
    xlsx_file = Folder + 'output_files/' + 'Linking_Tool.xlsx' # path to the output file
    xlsx_file_run = Folder + 'run_files/Run_file.xlsx' # path to the run file
    
    # list of the technologies of the waste sector for energy outputs
    sheet_obj_in = wb_obj_in['DO NOT CHANGE']
    list_tech_Energy_benchmark_WAS = []
    tech_Energy_benchmark_WAS = sheet_obj_in.cell(2 , 3)
    j=0
    while tech_Energy_benchmark_WAS.value != None:
        list_tech_Energy_benchmark_WAS.append(tech_Energy_benchmark_WAS.value)
        j+=1
        tech_Energy_benchmark_WAS = sheet_obj_in.cell(2 + j , 3)
        if j>100:
            break
    strlist_tech_Energy_benchmark_WAS = "('" + "' , '".join([str(tech) for tech in list_tech_Energy_benchmark_WAS]) + "')"

    # connection to the database specified in the input file
    sheet_obj_in = wb_obj_in['Input']
    db_file = sheet_obj_in.cell(2 , 3).value
    db_path = Folder + 'data_files/' + db_file + '.sqlite'
    conn = create_connection(db_path)
    cur = conn.cursor()

    print('Starting and open database')


    # # ------------------------------------------------------------------------ Create new table with technologies' nomenclature in the database ------------------------------------------------------------------------
    drop_table_if_exists = 'DROP TABLE IF EXISTS technology_nomenclature_ENE'
    cur.execute(drop_table_if_exists)
    create_table = 'CREATE TABLE technology_nomenclature_ENE(Sub TEXT NOT NULL, Sector TEXT NOT NULL, EndUse TEXT, Energy_generated TEXT, Vintage_Availability_year INTEGER, Tech_cat TEXT, Tech_type_1 TEXT, Tech_type_2 TEXT, Tech_type_3 TEXT, Energy_cat TEXT, Label_tech TEXT, ID_TEMOA_TO TEXT NOT NULL, Description_TEMOA_TO TEXT NOT NULL, Subsector TEXT, Capacity_units TEXT, Output_units TEXT);'
    file = open(Folder + "run_files/ENE/ENE_technology_nomenclature.csv")
    insert_records = 'INSERT INTO technology_nomenclature_ENE (Sub , Sector , EndUse , Energy_generated , Vintage_Availability_year , Tech_cat , Tech_type_1 , Tech_type_2 , Tech_type_3 , Energy_cat , Label_tech , ID_TEMOA_TO , Description_TEMOA_TO , Subsector, Capacity_Units, Output_Units) VALUES(? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? )'
    cur.execute(create_table)
    contents = csv.reader(file)
    cur.executemany(insert_records, contents)
    conn.commit()

    print('Importing technologies nomenclature to database')


    # # ------------------------------------------------------------------------ Extrat data from database ------------------------------------------------------------------------
    dict_df = {}
    # create queries
    # Import Electricity (grid)
    query = "SELECT t_periods , t_season , t_day , SUM(vflow_in) as 'electricity imported (TJ)' , SUBSTR(tech , 1 , 6) as subsector " + \
            "FROM Output_VFlow_In " + \
            "WHERE sector = 'TRAENE' AND input_comm = 'IMPELC' " + \
            "GROUP BY t_periods , t_season , t_day , subsector"
    dict_df['dfimp'] = query

    # PV and Wind capacities
    query = "SELECT t_periods , SUM(capacity) as 'capacity (MW)' , SUBSTR(tech, 13 , 2) as type " + \
            "FROM Output_CapacityByPeriodAndTech " + \
            "WHERE sector = 'ELCENE' AND ( tech LIKE '%PV%' OR tech LIKE '%WT%' ) " + \
            "GROUP BY t_periods , type "
    dict_df['dfPVnWT'] = query

    # Residential Storage
    query = "SELECT t_periods , SUM(capacity) as 'capacity (MW)' " + \
            "FROM Output_CapacityByPeriodAndTech " + \
            "WHERE sector = 'ELCENE' AND tech LIKE '%STG%' " + \
            "GROUP BY t_periods"
    dict_df['dfbatcap'] = query

    # Turbine & Engine (CHP)
    query = "SELECT t_periods , t_season , SUM(vflow_out) as 'electricity produced (TJ)' , SUBSTR(tech , 13 , 3) as 'type' " + \
            "FROM Output_VFlow_Out " + \
            "WHERE sector = 'DISENE' AND ( tech LIKE '%GTU%' OR tech LIKE '%ICE%' ) AND output_comm LIKE '%BDGELC' " + \
            "GROUP BY  t_periods , t_season , type"
    dict_df['dfEnTout'] = query

    query = "SELECT t_periods , SUM(capacity) as 'capacity (MW)' , SUBSTR(tech , 13 , 3) as 'type' " + \
            "FROM Output_CapacityByPeriodAndTech " + \
            "WHERE sector = 'DISENE' AND ( tech LIKE '%GTU%' OR tech LIKE '%ICE%' ) " + \
            "GROUP BY t_periods , type"
    dict_df['dfEnTcap'] = query

    # Electric vehicule charging
    query = "SELECT t_periods , t_season , SUM(vflow_in) as 'electricity consumed (TJ)' " + \
            "FROM Output_VFlow_In " + \
            "WHERE input_comm LIKE '%TRAELC' and output_comm LIKE '%TRABELC%' " + \
            "GROUP BY  t_periods , t_season"
    dict_df['dfEVin'] = query

    query = "SELECT Output_VFlow_In.t_periods , SUM(Output_VFlow_In.vflow_in) as 'electricity consumed (TJ)' , CASE WHEN technology_nomenclature_ENE.Tech_type_1 = 'RES' THEN 'RES' ELSE 'NON RES' END AS type " + \
            "FROM Output_VFlow_In , technology_nomenclature_ENE " + \
            "WHERE Output_VFlow_In.sector = 'STAENE' AND technology_nomenclature_ENE.ID_TEMOA_TO = Output_VFlow_In.tech " + \
            "GROUP BY  t_periods , type"
    dict_df['dfEVintype'] = query

    # Total electricity consumed
    query = "SELECT t_periods , t_season , SUM(vflow_in) as 'VF_IN_WASnTRA' " + \
            "FROM Output_VFlow_In " + \
            "WHERE ( input_comm LIKE '%WASELC' AND tech IN " + strlist_tech_Energy_benchmark_WAS + " ) OR (sector LIKE '%TRA' AND input_comm LIKE '%ELC%' AND input_comm NOT LIKE '%BELC%') " + \
            "GROUP BY t_periods , t_season"
    dict_df['dfWnTout'] = query

    query = "SELECT t_periods , t_season , SUM(vflow_in) as 'VF_IN_BDG' " + \
            "FROM Output_VFlow_In " + \
            "WHERE sector <> 'ELCENE' AND input_comm LIKE '%BDGELC' " + \
            "GROUP BY t_periods , t_season"
    dict_df['dfBDGin'] = query

    query = "SELECT t_periods , t_season , SUM(vflow_in) as 'VF_IN_ELCENE' " + \
            "FROM Output_VFlow_in " + \
            "WHERE input_comm LIKE '%BDGELC' AND output_comm LIKE '%BDGELC' " + \
            "GROUP BY t_periods , t_season"
    dict_df['dfELCENEin'] = query

    query = "SELECT t_periods , t_season , SUM(vflow_out) as 'VF_OUT_ELCENE' " + \
            "FROM Output_VFlow_Out " + \
            "WHERE sector = 'ELCENE' " + \
            "GROUP BY t_periods , t_season"
    dict_df['dfELCENEout'] = query

    query = "SELECT t_periods , t_season , SUM(vflow_out) as 'VF_OUT_DISENE' " + \
            "FROM Output_VFlow_Out " + \
            "WHERE sector = 'DISENE' AND output_comm LIKE '%BDGELC' " + \
            "GROUP BY t_periods , t_season "
    dict_df['dfDISENEout'] = query

    print('Extracting data from database')

    # Execute queries  
    i = 0
    for key in dict_df:
        cur.execute(dict_df[key]) # execute query
        columns = [column[0] for column in cur.description] # columns names
        data = [] # creation of an empty list that will be filled with query's output data
        rows = cur.fetchall()
        for row in rows: # filling of the list as dictonnary object
            data.append(dict(zip(columns, row)))
        dict_df[key] = pd.DataFrame.from_dict(data) # creation of a data frame with query's outputs
        i += 1
    conn.close()

    # # ------------------------------------------------------------------------ Add dataframes to excel run file ------------------------------------------------------------------------
    with pd.ExcelWriter(xlsx_file_run , mode = 'a' , engine = 'openpyxl' , if_sheet_exists = 'replace') as writer :
        dict_df['dfimp'].to_excel(writer , 'Import Electricity')
        dict_df['dfPVnWT'].to_excel(writer , 'PV & WIND Capacity')
        dict_df['dfbatcap'].to_excel(writer , 'Residential Storage Capacity')
        dict_df['dfEnTcap'].to_excel(writer , 'Turbine & Engine Cap and Energy')
        dict_df['dfEVin'].to_excel(writer , 'Elec Veh Energy')
        dict_df['dfBDGin'].to_excel(writer , 'Elec Energy Consumed')
        dict_df['dfEVintype'].to_excel(writer , 'Elec Veh Energy per type')

    with pd.ExcelWriter(xlsx_file_run , mode = 'a' , engine = 'openpyxl' , if_sheet_exists = 'overlay') as writer :
        dict_df['dfELCENEout'].to_excel(writer , 'Elec Energy Consumed' , startcol = 12)
        dict_df['dfELCENEin'].to_excel(writer , 'Elec Energy Consumed' , startcol = 18)
        dict_df['dfDISENEout'].to_excel(writer , 'Elec Energy Consumed' , startcol = 24)
        dict_df['dfWnTout'].to_excel(writer , 'Elec Energy Consumed' , startcol = 6)
        dict_df['dfEnTout'].to_excel(writer , 'Turbine & Engine Cap and Energy' , startcol = writer.sheets['Turbine & Engine Cap and Energy'].max_column + 1)

    print('Writing data into the run file')


    # # ------------------------------------------------------------------------ Open and close xlsx file to save modifications ------------------------------------------------------------------------
    app = xw.App(visible=False)
    book = xw.Book(xlsx_file_run)
    book.save()
    book.close()
    app.quit()

    print('Saving the run file')


    # # ------------------------------------------------------------------------ Creation of the matrices for the linking tool (=spatial disaggregation) ------------------------------------------------------------------------
    wb_obj_run = openpyxl.load_workbook(xlsx_file_run, data_only = True) # open the run file
    list_year = dict_df['dfimp']['t_periods'].unique().tolist() # list of years modelled in TEMOA
    list_year.sort()
    nb_of_year = len(list_year)
    nb_of_yearseason = nb_of_year*4

    # Energy consumed matrix
    sheet_obj_run = wb_obj_run['DO NOT CHANGE']
    mat_elec_consumed = [[sheet_obj_run.cell(2 + r , 8 + c).value for c in range(nb_of_year)] for r in range(nb_of_yearseason)]
    list_yearseason = [sheet_obj_run.cell(2 + r , 7).value for r in range(nb_of_yearseason)]

    sheet_obj_in = wb_obj_in['POP_TRZ_%']
    mat_pop_TRZ = [[sheet_obj_in.cell(2 + r , 2 + c).value for r in range(650)] for c in range(nb_of_year)]

    sheet_obj_in = wb_obj_in['Overlap_TS_TRZ_%']
    mat_TS_TRZ = [[sheet_obj_in.cell(2 + r , 2 + c).value for c in range(39)] for r in range(650)]
    list_TS = [sheet_obj_in.cell(1 , 2 + c).value for c in range(39)]

    mat_elec_per_TRZ = np.matmul(mat_elec_consumed , mat_pop_TRZ)
    mat_elec_per_TS = np.matmul(mat_elec_per_TRZ , mat_TS_TRZ)

    # Adjustment between actual transformer service areas and those modelled in SILVER-Toronto (within Toronto city, a few areas are linked with transformer stations outsides of city border and those stations aren't modeled in SILVER-Toronto)
    df_elec_conso = pd.DataFrame(mat_elec_per_TS , list_yearseason , list_TS)
    df_elec_conso['Rexdale'] = df_elec_conso['Rexdale'] - df_elec_conso['Woodbridge']
    df_elec_conso['Bridgman'] = df_elec_conso['Bridgman'] + df_elec_conso['High level']
    df_elec_conso['John'] = 0.5*df_elec_conso['Windsor']
    df_elec_conso['Copeland'] = 0.5*df_elec_conso['Windsor']
    df_elec_conso = df_elec_conso.drop(['Windsor'] , axis = 1)
    df_elec_conso = df_elec_conso.drop(['High level'] , axis = 1)
    df_elec_conso = df_elec_conso.drop(['Woodbridge'] , axis = 1)
    df_elec_consoMWh = df_elec_conso * 277.7778 # convert TJ to MWh
    df_elec_conso.insert(0, "season", ['F' , 'R' , 'S' , 'W']*nb_of_year, True)
    df_elec_consoMWh.insert(0, "season", ['F' , 'R' , 'S' , 'W']*nb_of_year, True) 
    df_elec_consoMWh.insert(0, "year", [x for item in list_year for x in itertools.repeat(item,4)], True)
    df_elec_conso.insert(0, "year", [x for item in list_year for x in itertools.repeat(item,4)], True)

    with pd.ExcelWriter(xlsx_file , mode = 'a' , engine = 'openpyxl' , if_sheet_exists = 'replace') as writer :
        df_elec_consoMWh.to_excel(writer , 'TLi')

    df_elec_conso[df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist()] = df_elec_conso[df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist()].apply(lambda x: x/x.sum() if x.sum() !=0 else 0, axis=1)

    print('Creating energy consumption matrix')

    # PV capacities matrix
    mat_capacity_PV = [[sheet_obj_run.cell(2 + r , 26 + c).value for c in range(nb_of_year)] for r in range(nb_of_year)]##*## (first column index)
    mat_capacity_PV_TS = np.matmul(mat_capacity_PV , df_elec_conso.drop_duplicates(subset = ['year']).drop(columns = {'year' , 'season'}).to_numpy())
    df13 = pd.DataFrame(mat_capacity_PV_TS , list_year , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())

    print('Creating PV capacities matrix')

    # open and close xlsx file to save modifications
    app = xw.App(visible=False)
    book = xw.Book(xlsx_file)
    book.save()
    book.close()
    app.quit()

    print('Saving linking tool output')

    # adjustment of PV capacities by transformer nodes
    wb = openpyxl.load_workbook(xlsx_file , data_only = True)
    sheet = wb['Si']
    mat_S = [[sheet.cell(3 , 2 + c).value for c in range(36)]]
    dfcapmax = pd.DataFrame(mat_S , columns=df13.columns.tolist())
    for year in list_year: # no more PV capacity installed than available
        over_cap , i = 0 , 0
        list_TS2 = df13.columns.tolist()
        random.shuffle(list_TS2)
        for TS in list_TS2:
            if df13[TS][year] > dfcapmax[TS][0]*0.95:
                over_cap += df13[TS][year] - dfcapmax[TS][0]*0.95
                df13[TS][year] = dfcapmax[TS][0]*0.95
            elif df13[TS][year] < dfcapmax[TS][0]*0.95:
                if over_cap > 0:
                    if over_cap < dfcapmax[TS][0]*0.95 - df13[TS][year]:
                        df13[TS][year] += over_cap
                        over_cap = 0
                    else:
                        over_cap += df13[TS][year] - dfcapmax[TS][0]*0.95
                        df13[TS][year] = dfcapmax[TS][0]*0.95
        while over_cap > 0:
            i += 1
            if i>10:
                break
            random.shuffle(list_TS2)
            for TS in list_TS2:
                if df13[TS][year] > dfcapmax[TS][0]*0.95:
                    over_cap += df13[TS][year] - dfcapmax[TS][0]*0.95
                    df13[TS][year] = dfcapmax[TS][0]*0.95
                elif df13[TS][year] < dfcapmax[TS][0]*0.95:
                    if over_cap > 0:
                        if over_cap < dfcapmax[TS][0]*0.95 - df13[TS][year]:
                            df13[TS][year] += over_cap
                            over_cap = 0
                        else:
                            over_cap += df13[TS][year] - dfcapmax[TS][0]*0.95
                            df13[TS][year] = dfcapmax[TS][0]*0.95
    df13['Total per year'] = df13.sum(axis = 1)

    print('Adjusting PV capacities by transformer nodes')

    # Wind capacities matrix
    mat_capacity_WT = [[sheet_obj_run.cell(19 + r , 26 + c).value for c in range(nb_of_year)] for r in range(nb_of_year)]##*## (first row and column indexes)
    mat_capacity_WT_TS = np.matmul(mat_capacity_WT , df_elec_conso.drop_duplicates(subset = ['year']).drop(columns = {'year' , 'season'}).to_numpy())
    df14 = pd.DataFrame(mat_capacity_WT_TS , list_year , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())
    df14['Total per year'] = df14.sum(axis = 1)

    print('Creating wind capacities matrix')

    # Residential storage capacities matrix
    mat_capacity_STG = [[sheet_obj_run.cell(2 + r , 44 + c).value for c in range(nb_of_year)] for r in range(nb_of_year)]##*## (first column index)
    mat_capacity_STG_TS = np.matmul(mat_capacity_STG , df_elec_conso.drop_duplicates(subset = ['year']).drop(columns = {'year' , 'season'}).to_numpy())
    dfbat = pd.DataFrame(mat_capacity_STG_TS , list_year , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())
    dfbat['Total per year'] = dfbat.sum(axis = 1)

    print('Creating residential storage matrix')

    # electric vehicles charging energy matrices
    mat_charging_EV = [[sheet_obj_run.cell(2 + r , 68 + c).value for c in range(nb_of_yearseason)] for r in range(nb_of_yearseason)]##*## (first column index)
    mat_charging_EV_TS = np.matmul(mat_charging_EV , df_elec_conso.drop(columns = {'year' , 'season'}).to_numpy())
    dfEV = pd.DataFrame(mat_charging_EV_TS , list_yearseason , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())
    dfEV.insert(0, "season", ['F' , 'R' , 'S' , 'W']*nb_of_year, True)
    dfEV.insert(0, "year", [x for item in list_year for x in itertools.repeat(item,4)], True)

    mat_charging_EV_RES = [[sheet_obj_run.cell(2 + r , 218 + c).value for c in range(nb_of_year)] for r in range(nb_of_year)]##*## (first column index)
    mat_charging_EV_RES_TS = np.matmul(mat_charging_EV_RES , df_elec_conso.drop_duplicates(subset = ['year']).drop(columns = {'year' , 'season'}).to_numpy())
    dfEVRES = pd.DataFrame(mat_charging_EV_RES_TS , list_year , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())

    print('Creating electric vehicles matrices')

    # engine and turbine (CHP) matrices : vflow out and capacity
    mat_capacity_ICEnGTU = [[sheet_obj_run.cell(2 + r , 131 + c).value for c in range(nb_of_year)] for r in range(nb_of_year)]##*## (first column index)
    mat_capacity_ICEnGTU_TS = np.matmul(mat_capacity_ICEnGTU , df_elec_conso.drop_duplicates(subset = ['year']).drop(columns = {'year' , 'season'}).to_numpy())
    dfICEnGTUcap = pd.DataFrame(mat_capacity_ICEnGTU_TS , list_year , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())

    mat_elec_produced_ICEnGTU = [[sheet_obj_run.cell(2 + r , 155 + c).value for c in range(nb_of_yearseason)] for r in range(nb_of_yearseason)]##*## (first column index)

    mat_elec_produced_ICEnGTU_TS = np.matmul(mat_elec_produced_ICEnGTU , df_elec_conso.drop(columns = {'year' , 'season'}).to_numpy())
    dfICEnGTU = pd.DataFrame(mat_elec_produced_ICEnGTU_TS , list_yearseason , df_elec_conso.drop(columns = {'year' , 'season'}).columns.tolist())
    dfICEnGTU.insert(0, "season", ['F' , 'R' , 'S' , 'W']*nb_of_year, True)
    dfICEnGTU.insert(0, "year", [x for item in list_year for x in itertools.repeat(item,4)], True)

    print('Creating CHP matrices')


    # # ------------------------------------------------------------------------ Add final dataframes to output excel spreadsheet ------------------------------------------------------------------------
    with pd.ExcelWriter(xlsx_file , mode = 'a' , engine = 'openpyxl' , if_sheet_exists = 'replace') as writer :
        df13.to_excel(writer , 'PVi')
        df14.to_excel(writer , 'Wi')
        dfbat.to_excel(writer , 'STi')
        dfEV.to_excel(writer , 'EVi')
        dfEVRES.to_excel(writer , 'EVri')
        dfICEnGTU.to_excel(writer , 'ETi')
        dfICEnGTUcap.to_excel(writer , 'ETci')

    print('Adding matrices to output linking tool file')


main()